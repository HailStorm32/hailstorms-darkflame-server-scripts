#!/usr/bin/env python3
"""Generate a spreadsheet report of account playtimes.

The script connects to the server database, gathers each account's creation
and most recent play information, then writes the results to an Excel file.
Rows for banned accounts are highlighted in red while locked accounts are
highlighted in yellow.  The report is sorted from longest play duration to
shortest and also includes a statistics sheet summarising the minimum, maximum
and average playtimes for older accounts.  A separate sheet lists active
accounts with at least one character, sorted by how long it has been since they
last played.
"""

from __future__ import annotations

import argparse
from datetime import datetime, timedelta, timezone
from typing import Optional, Tuple, List, Dict

try:
    from playerCntSettings import (
        DATABASE_IP,
        DATABASE_NAME,
        DATABASE_USER,
        DATABASE_PASS,
    )
except Exception as exc:  # pragma: no cover - dependency may be missing
    raise SystemExit("playerCntSettings.py is required: {}".format(exc))

try:
    import pymysql  # type: ignore
except Exception as exc:  # pragma: no cover - dependency may be missing
    raise SystemExit("pymysql is required: {}".format(exc))

try:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import PatternFill  # type: ignore
except Exception as exc:  # pragma: no cover - dependency may be missing
    raise SystemExit("openpyxl is required: {}".format(exc))


# ---------------------------------------------------------------------------
# Database access
# ---------------------------------------------------------------------------


# MySQL connection settings sourced from playerCntSettings.py
DB_PORT = 3306

# Minimum age (in days) an account must be before it counts toward statistics
STATS_MIN_ACCOUNT_AGE_DAYS = 1


def get_connection():
    """Create a MySQL database connection using the configured credentials."""
    return pymysql.connect(
        host=DATABASE_IP,
        user=DATABASE_USER,
        password=DATABASE_PASS,
        database=DATABASE_NAME,
        port=DB_PORT,
    )


# ---------------------------------------------------------------------------
# Query helpers
# ---------------------------------------------------------------------------


def get_accounts(cur) -> list:
    """Return list of account records."""
    cur.execute("SELECT id, name, banned, locked, created_at FROM accounts;")
    return cur.fetchall()


def get_last_character(cur, account_id: int) -> Tuple[Optional[str], Optional[int]]:
    """Return the latest character name and last login time for account."""
    cur.execute(
        f"SELECT name, last_login FROM charinfo WHERE account_id = {account_id} ORDER BY last_login DESC LIMIT 1;"
    )
    row = cur.fetchone()
    if not row:
        return None, None
    return row[0], row[1]


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")


def compute_stats(deltas: List[timedelta]) -> Dict[str, Optional[timedelta]]:
    """Return min, max and average from a list of timedeltas."""
    if not deltas:
        return {"min": None, "max": None, "avg": None}
    min_delta = min(deltas)
    max_delta = max(deltas)
    avg_seconds = sum(d.total_seconds() for d in deltas) / len(deltas)
    avg_delta = timedelta(seconds=avg_seconds)
    return {"min": min_delta, "max": max_delta, "avg": avg_delta}


def write_report(
    rows: list,
    stats_all: Dict[str, Optional[timedelta]],
    stats_locked: Dict[str, Optional[timedelta]],
    age_rows: list,
    output: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Playtimes"
    ws.append(
        [
            "Account Name",
            "Last Character",
            "Account Creation Date",
            "Last Played Date",
            "Time From Creation To Last Played",
            "Time Since Last Played",
            "Status",
        ]
    )

    for r in rows:
        ws.append(r[:7])
        idx = ws.max_row
        fill = r[7]
        if fill:
            for cell in ws[idx]:
                cell.fill = fill

    stats_ws = wb.create_sheet(title="Statistics")
    stats_ws.append(["Metric", "All Accounts", "Locked Accounts"])
    stats_ws.append(
        [
            "Min",
            str(stats_all["min"]) if stats_all["min"] else "",
            str(stats_locked["min"]) if stats_locked["min"] else "",
        ]
    )
    stats_ws.append(
        [
            "Max",
            str(stats_all["max"]) if stats_all["max"] else "",
            str(stats_locked["max"]) if stats_locked["max"] else "",
        ]
    )
    stats_ws.append(
        [
            "Average",
            str(stats_all["avg"]) if stats_all["avg"] else "",
            str(stats_locked["avg"]) if stats_locked["avg"] else "",
        ]
    )

    age_ws = wb.create_sheet(title="Accounts By Age")
    age_ws.append(
        [
            "Account Name",
            "Last Character",
            "Account Creation Date",
            "Last Played Date",
            "Time From Creation To Last Played",
            "Time Since Last Played",
            "Status",
        ]
    )
    for r in age_rows:
        age_ws.append(r[:7])

    wb.save(output)


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate account playtime report")
    parser.add_argument("--output", default="account_playtimes.xlsx", help="Output xlsx filename")
    args = parser.parse_args()

    conn = get_connection()
    cur = conn.cursor()
    accounts = get_accounts(cur)

    rows = []
    stats_deltas: List[timedelta] = []
    locked_stats_deltas: List[timedelta] = []
    now = datetime.utcnow()
    min_age = timedelta(days=STATS_MIN_ACCOUNT_AGE_DAYS)

    for acc in accounts:
        acc_id, name, banned, locked, created_at = acc
        char_name, last_login = get_last_character(cur, acc_id)

        if isinstance(created_at, str):
            created_dt = datetime.fromisoformat(created_at)
        else:
            created_dt = created_at

        last_login_dt = (
            datetime.fromtimestamp(last_login, tz=timezone.utc).replace(tzinfo=None)
            if last_login
            else None
        )

        delta: Optional[timedelta] = None
        if last_login_dt:
            delta = last_login_dt - created_dt

        status = "Banned" if banned else "Locked" if locked else "Active"
        fill = RED_FILL if banned else YELLOW_FILL if locked else None

        age = now - created_dt
        if delta and age >= min_age:
            stats_deltas.append(delta)
            if locked:
                locked_stats_deltas.append(delta)

        since_last_login = now - last_login_dt if last_login_dt else None

        rows.append(
            [
                name,
                char_name or "",
                created_dt.isoformat(sep=" "),
                last_login_dt.isoformat(sep=" ") if last_login_dt else "",
                str(delta) if delta else "",
                str(since_last_login) if since_last_login else "",
                status,
                fill,
                delta,
                age,
                last_login_dt,
            ]
        )

    rows.sort(key=lambda r: r[8] or timedelta(0), reverse=True)
    age_rows = sorted(
        [r for r in rows if r[6] == "Active" and r[1]],
        key=lambda r: r[10] or datetime.min,
    )
    stats_all = compute_stats(stats_deltas)
    stats_locked = compute_stats(locked_stats_deltas)
    write_report(rows, stats_all, stats_locked, age_rows, args.output)
    print(f"Wrote report to {args.output}")


if __name__ == "__main__":  # pragma: no cover
    main()